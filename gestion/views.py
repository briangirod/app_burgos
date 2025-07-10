from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from .models import Registro, ProductoUtilizado, Producto, Tratamiento, Consultorio, PacienteRecepcion, EstadoDoctor, Sector, PerfilUsuario
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from django.db.models import Q
from django.utils.dateparse import parse_date
from django.contrib.auth.models import User, Group
from django.http import JsonResponse, HttpResponse
import json
from django.template.loader import render_to_string
from django.utils.timezone import now, localdate
from django.http import HttpResponseForbidden
from django.shortcuts import get_object_or_404
from django.views.decorators.http import require_POST
from django.db import transaction, IntegrityError
from django.core.cache import cache
from django.contrib.auth import authenticate, update_session_auth_hash
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

logger = logging.getLogger(__name__)

def obtener_nombre_completo_usuario(user):
    """
    Obtiene el nombre completo del usuario desde su perfil o fallback al username
    """
    try:
        perfil = PerfilUsuario.objects.get(usuario=user)
        return perfil.nombre_completo()
    except PerfilUsuario.DoesNotExist:
        return user.get_full_name() if user.get_full_name() else user.username

@login_required
def home(request):
    # Verificar si el usuario pertenece a algún grupo de caja
    if request.user.groups.filter(name__in=['Caja', 'Caja_1', 'Caja_2']).exists():
        return redirect('panel_caja')
    else:
        return redirect('panel_doctor')

def get_or_create_safe(model_class, **kwargs):
    """
    Versión segura de get_or_create que maneja race conditions
    """
    try:
        return model_class.objects.get(**kwargs), False
    except model_class.DoesNotExist:
        try:
            with transaction.atomic():
                return model_class.objects.create(**kwargs), True
        except IntegrityError:
            # Otro proceso ya creó el objeto, intentar obtenerlo nuevamente
            return model_class.objects.get(**kwargs), False

@login_required
@transaction.atomic
def panel_doctor(request):
    if request.method == 'POST':
        try:
            nombre_paciente = request.POST.get('nombre_paciente')
            tratamiento_nombre = request.POST.get('tratamiento_nombre')
            consultorio_nombre = request.POST.get('consultorio_nombre')
            caja_destino_id = request.POST.get('caja_destino')

            # Validaciones
            if not all([nombre_paciente, tratamiento_nombre, consultorio_nombre, caja_destino_id]):
                messages.error(request, "Todos los campos son obligatorios.")
                return redirect('panel_doctor')
            
            # Buscar o crear Tratamiento (thread-safe)
            tratamiento, _ = get_or_create_safe(Tratamiento, nombre=tratamiento_nombre)

            # Buscar o crear Consultorio (thread-safe)
            consultorio, _ = get_or_create_safe(Consultorio, nombre=consultorio_nombre)

            # Obtener el sector/caja destino
            try:
                caja_destino = Sector.objects.get(id=caja_destino_id)
            except Sector.DoesNotExist:
                messages.error(request, "Sector de destino inválido.")
                return redirect('panel_doctor')

            # Crear Registro
            registro = Registro.objects.create(
                nombre_paciente=nombre_paciente,
                tratamiento=tratamiento,
                doctor=request.user,
                consultorio=consultorio,
                caja_destino=caja_destino
            )

            # Procesar productos y cantidades
            producto_nombres = request.POST.getlist('producto_nombre')
            cantidades = request.POST.getlist('cantidad')

            productos_creados = []
            for nombre, cantidad in zip(producto_nombres, cantidades):
                if not nombre.strip() or not cantidad.strip():
                    continue  # Ignorar productos vacíos

                try:
                    cantidad_int = int(cantidad)
                    if cantidad_int <= 0:
                        continue
                except (ValueError, TypeError):
                    continue

                # Buscar o crear Producto (thread-safe)
                producto, _ = get_or_create_safe(Producto, nombre=nombre.strip())
                
                ProductoUtilizado.objects.create(
                    registro=registro,
                    producto=producto,
                    cantidad=cantidad_int
                )
                productos_creados.append(f"{producto.nombre} x{cantidad_int}")

            # Invalidar cache de caja
            try:
                cache.clear()
            except:
                pass  # Si falla el cache, no es crítico
            
            if productos_creados:
                messages.success(request, f"Registro enviado correctamente a {caja_destino.nombre}. Productos: {', '.join(productos_creados)}")
            else:
                messages.warning(request, "Registro creado sin productos.")
                
            logger.info(f"Doctor {request.user.username} creó registro {registro.id} para {caja_destino.nombre}")
            
        except Exception as e:
            logger.error(f"Error en panel_doctor: {str(e)}")
            messages.error(request, "Error al procesar el registro. Intente nuevamente.")
            return redirect('panel_doctor')
            
        return redirect('panel_doctor')

    # GET request - mostrar formulario
    productos = Producto.objects.all().order_by('nombre')
    tratamientos = Tratamiento.objects.all().order_by('nombre')
    consultorios = Consultorio.objects.all().order_by('nombre')
    sectores = Sector.objects.all().order_by('nombre')
    
    return render(request, 'gestion/doctor_panel.html', {
        'productos': productos,
        'tratamientos': tratamientos,
        'consultorios': consultorios,
        'sectores': sectores,
    })
@login_required
def panel_caja(request):
    hoy = timezone.now().date()
    
    # Determinar qué registros puede ver este usuario basado en su grupo
    registros = None
    
    # Verificar grupos del usuario
    user_groups = request.user.groups.values_list('name', flat=True)
    
    if 'Caja_1' in user_groups:
        # Usuario del grupo Caja_1 ve solo registros enviados a "Caja 1"
        try:
            sector_caja1 = Sector.objects.get(nombre__icontains='Caja 1')
            registros = Registro.objects.filter(
                fecha=hoy, 
                estado='para_abonar',
                caja_destino=sector_caja1
            )
        except Sector.DoesNotExist:
            registros = Registro.objects.none()
            
    elif 'Caja_2' in user_groups:
        # Usuario del grupo Caja_2 ve solo registros enviados a "Caja 2"
        try:
            sector_caja2 = Sector.objects.get(nombre__icontains='Caja 2')
            registros = Registro.objects.filter(
                fecha=hoy, 
                estado='para_abonar',
                caja_destino=sector_caja2
            )
        except Sector.DoesNotExist:
            registros = Registro.objects.none()
            
    elif 'Caja' in user_groups:
        # Usuario del grupo Caja genérico ve todos los registros
        registros = Registro.objects.filter(fecha=hoy, estado='para_abonar')
        
    else:
        # Fallback: intentar usar el perfil de usuario
        try:
            perfil_usuario = PerfilUsuario.objects.get(usuario=request.user)
            sector_usuario = perfil_usuario.sector
            registros = Registro.objects.filter(
                fecha=hoy, 
                estado='para_abonar',
                caja_destino=sector_usuario
            )
        except PerfilUsuario.DoesNotExist:
            # Si no tiene perfil ni grupos apropiados, no mostrar nada
            registros = Registro.objects.none()

    if request.method == 'POST':
        registro_id = request.POST.get('registro_id')
        Registro.objects.filter(id=registro_id).update(estado='abonado')
        return redirect('panel_caja')

    return render(request, 'gestion/caja_panel.html', {'registros': registros})


@login_required
def historial_caja(request):
    registros = Registro.objects.filter(estado='abonado').order_by('-fecha')

    producto_nombre = request.GET.get('producto')
    doctor_id = request.GET.get('doctor')
    fecha_inicio = request.GET.get('desde')
    fecha_fin = request.GET.get('hasta')
    caja_id = request.GET.get('caja')

    # Aplicar filtros
    if producto_nombre:
        registros = registros.filter(productoutilizado__producto__nombre=producto_nombre)
    if doctor_id:
        registros = registros.filter(doctor__id=doctor_id)
    if fecha_inicio:
        registros = registros.filter(fecha__gte=parse_date(fecha_inicio))
    if fecha_fin:
        registros = registros.filter(fecha__lte=parse_date(fecha_fin))
    if caja_id:
        registros = registros.filter(caja_destino__id=caja_id)

    # Datos para los filtros
    productos = Producto.objects.all().order_by('nombre')
    grupo_doctores = Group.objects.get(name="Doctor")  # Asegurate de tener este grupo
    doctores = User.objects.filter(groups=grupo_doctores)
    cajas = Sector.objects.all().order_by('nombre')

    # Calcular total de registros
    total_registros = registros.count()
    
    return render(request, 'gestion/historial.html', {
        'registros': registros,
        'productos': productos,
        'doctores': doctores,
        'cajas': cajas,
        'total_registros': total_registros,
        'producto_filtro': producto_nombre or '',
        'doctor_filtro': int(doctor_id) if doctor_id else '',
        'caja_filtro': int(caja_id) if caja_id else '',
        'desde': fecha_inicio or '',
        'hasta': fecha_fin or '',
    })

@login_required
def historial_doctor(request):
    registros = Registro.objects.filter(doctor=request.user).order_by('-fecha')
    return render(request, 'gestion/doctor_historial.html', {'registros': registros})


@login_required
def caja_panel_ajax(request):
    try:
        # Cache key único por usuario y fecha
        fecha_actual = localdate()
        cache_key = f"caja_panel_{request.user.id}_{fecha_actual}"
        
        # Intentar obtener desde cache
        cached_html = cache.get(cache_key)
        if cached_html:
            return JsonResponse({'html': cached_html})

        # Determinar qué registros puede ver este usuario basado en su grupo
        user_groups = request.user.groups.values_list('name', flat=True)
        registros = None
        
        if 'Caja_1' in user_groups:
            # Usuario del grupo Caja_1 ve solo registros enviados a "Caja 1"
            try:
                sector_caja1 = Sector.objects.get(nombre__icontains='Caja 1')
                registros = Registro.objects.filter(
                    fecha=fecha_actual, 
                    estado='para_abonar',
                    caja_destino=sector_caja1
                ).select_related(
                    'tratamiento', 
                    'doctor', 
                    'consultorio', 
                    'caja_destino'
                ).prefetch_related(
                    'productoutilizado_set__producto'
                ).order_by('-id')
            except Sector.DoesNotExist:
                registros = Registro.objects.none()
                
        elif 'Caja_2' in user_groups:
            # Usuario del grupo Caja_2 ve solo registros enviados a "Caja 2"
            try:
                sector_caja2 = Sector.objects.get(nombre__icontains='Caja 2')
                registros = Registro.objects.filter(
                    fecha=fecha_actual, 
                    estado='para_abonar',
                    caja_destino=sector_caja2
                ).select_related(
                    'tratamiento', 
                    'doctor', 
                    'consultorio', 
                    'caja_destino'
                ).prefetch_related(
                    'productoutilizado_set__producto'
                ).order_by('-id')
            except Sector.DoesNotExist:
                registros = Registro.objects.none()
                
        elif 'Caja' in user_groups:
            # Usuario del grupo Caja genérico ve todos los registros
            registros = Registro.objects.filter(
                fecha=fecha_actual, 
                estado='para_abonar'
            ).select_related(
                'tratamiento', 
                'doctor', 
                'consultorio', 
                'caja_destino'
            ).prefetch_related(
                'productoutilizado_set__producto'
            ).order_by('-id')
            
        else:
            # Fallback: intentar usar el perfil de usuario
            try:
                perfil_usuario = PerfilUsuario.objects.select_related('sector').get(usuario=request.user)
                sector_usuario = perfil_usuario.sector
                registros = Registro.objects.filter(
                    fecha=fecha_actual, 
                    estado='para_abonar',
                    caja_destino=sector_usuario
                ).select_related(
                    'tratamiento', 
                    'doctor', 
                    'consultorio', 
                    'caja_destino'
                ).prefetch_related(
                    'productoutilizado_set__producto'
                ).order_by('-id')
            except PerfilUsuario.DoesNotExist:
                # Fallback para compatibilidad
                registros = Registro.objects.filter(
                    fecha=fecha_actual, 
                    estado='para_abonar'
                ).select_related(
                    'tratamiento', 
                    'doctor', 
                    'consultorio', 
                    'caja_destino'
                ).prefetch_related(
                    'productoutilizado_set__producto'
                ).order_by('-id')

        html = render_to_string('gestion/_tabla_panel_caja.html', {'registros': registros}, request=request)
        
        # Cachear por 30 segundos (menos que el intervalo de refresh)
        cache.set(cache_key, html, 30)
        
        return JsonResponse({'html': html})
        
    except Exception as e:
        logger.error(f"Error en caja_panel_ajax: {str(e)}")
        return JsonResponse({'error': 'Error interno del servidor'}, status=500)
    
@login_required
def panel_recepcion(request):
    if not request.user.groups.filter(name='Recepcion').exists():
        return HttpResponseForbidden("No tenés permiso para acceder a este panel.")
    return render(request, 'gestion/recepcion_panel.html')

@login_required
def estado_pacientes_ajax(request):
    try:
        # Cache key para estado de pacientes
        cache_key = f"estado_pacientes_{localdate()}"
        cached_html = cache.get(cache_key)
        if cached_html:
            return HttpResponse(cached_html)

        # Optimizar consultas con select_related
        pacientes = PacienteRecepcion.objects.select_related('doctor_asignado').order_by('-fecha_ingreso')[:50]  # Limitar resultados

        # Optimizar consulta de doctores con su estado
        doctores = User.objects.filter(
            groups__name="Doctor"
        ).select_related('estado_doctor').order_by('username')

        # Preparar datos de doctores sin N+1 queries
        for doctor in doctores:
            if hasattr(doctor, 'estado_doctor') and doctor.estado_doctor:
                doctor.estado_display = doctor.estado_doctor.get_estado_display()
                doctor.estado_clase = doctor.estado_doctor.estado_clase()
            else:
                doctor.estado_display = 'Desconectado'
                doctor.estado_clase = 'estado-desconectado'

        html = render_to_string('gestion/_tabla_pacientes.html', {
            'pacientes': pacientes,
            'doctores': doctores
        })

        # Cachear por 10 segundos
        cache.set(cache_key, html, 10)
        
        return HttpResponse(html)
        
    except Exception as e:
        logger.error(f"Error en estado_pacientes_ajax: {str(e)}")
        return HttpResponse('<tr><td colspan="5">Error al cargar datos</td></tr>')




@csrf_exempt
@login_required
def ingresar_paciente_ajax(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        PacienteRecepcion.objects.create(
            nombre=data.get('nombre_apellido'),
            dni=data.get('dni', ''),
            fecha_nacimiento=data.get('fecha_nacimiento') or None,
            telefono=data.get('telefono', ''),
        )
        return JsonResponse({'status': 'ok'})
    return JsonResponse({'error': 'Método no permitido'}, status=405)


@require_POST
@login_required
def eliminar_paciente(request, id):
    paciente = get_object_or_404(PacienteRecepcion, id=id)
    paciente.delete()
    return HttpResponse(status=204)

# FUNCIÓN COMENTADA PARA VERSIÓN FUTURA - Estado del doctor
"""
@require_POST
@login_required
def actualizar_estado_doctor(request):
    try:
        data = json.loads(request.body)
        nuevo_estado = data.get('estado')

        if nuevo_estado not in dict(EstadoDoctor.ESTADOS):
            return JsonResponse({'error': 'Estado inválido'}, status=400)

        # Usar get_or_create_safe para evitar race conditions
        estado, _ = get_or_create_safe(EstadoDoctor, doctor=request.user)
        estado.estado = nuevo_estado
        estado.save()

        # Invalidar cache de estado de pacientes
        cache.delete_pattern(f"estado_pacientes_*")
        
        logger.info(f"Doctor {request.user.username} cambió estado a {nuevo_estado}")
        return JsonResponse({'success': True})
        
    except Exception as e:
        logger.error(f"Error en actualizar_estado_doctor: {str(e)}")
        return JsonResponse({'error': 'Error interno del servidor'}, status=500)
"""

@login_required
def perfil_usuario(request):
    """
    Vista para gestionar el perfil del usuario (nombre y apellido)
    """
    try:
        perfil = PerfilUsuario.objects.get(usuario=request.user)
    except PerfilUsuario.DoesNotExist:
        # Si no existe perfil, crear uno básico
        perfil = PerfilUsuario.objects.create(
            usuario=request.user,
            sector_id=1,  # Sector por defecto
            nombre='',
            apellido=''
        )
    
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        apellido = request.POST.get('apellido', '').strip()
        
        perfil.nombre = nombre
        perfil.apellido = apellido
        perfil.save()
        
        messages.success(request, 'Perfil actualizado correctamente.')
        return redirect('perfil_usuario')
    
    return render(request, 'gestion/perfil_usuario.html', {'perfil': perfil})

@login_required
def cambiar_contrasena(request):
    """
    Vista personalizada para cambio de contraseña
    """
    if request.method == 'POST':
        contrasena_actual = request.POST.get('current_password', '').strip()
        nueva_contrasena = request.POST.get('new_password', '').strip()
        confirmar_contrasena = request.POST.get('confirm_password', '').strip()
        
        # Validaciones
        if not contrasena_actual:
            messages.error(request, 'Debe ingresar su contraseña actual.')
            return redirect('cambiar_contrasena')
        
        if not nueva_contrasena:
            messages.error(request, 'Debe ingresar una nueva contraseña.')
            return redirect('cambiar_contrasena')
        
        if len(nueva_contrasena) < 8:
            messages.error(request, 'La nueva contraseña debe tener al menos 8 caracteres.')
            return redirect('cambiar_contrasena')
        
        if nueva_contrasena != confirmar_contrasena:
            messages.error(request, 'Las contraseñas nuevas no coinciden.')
            return redirect('cambiar_contrasena')
        
        # Verificar contraseña actual
        if not request.user.check_password(contrasena_actual):
            messages.error(request, 'La contraseña actual es incorrecta.')
            return redirect('cambiar_contrasena')
        
        # Verificar que la nueva contraseña sea diferente
        if request.user.check_password(nueva_contrasena):
            messages.error(request, 'La nueva contraseña debe ser diferente a la actual.')
            return redirect('cambiar_contrasena')
        
        try:
            # Cambiar contraseña
            request.user.set_password(nueva_contrasena)
            request.user.save()
            
            # Mantener la sesión activa después del cambio
            update_session_auth_hash(request, request.user)
            
            messages.success(request, 'Contraseña cambiada exitosamente.')
            logger.info(f"Usuario {request.user.username} cambió su contraseña")
            
            return redirect('home')
            
        except Exception as e:
            logger.error(f"Error al cambiar contraseña para {request.user.username}: {str(e)}")
            messages.error(request, 'Error al cambiar la contraseña. Intente nuevamente.')
            return redirect('cambiar_contrasena')
    
    return render(request, 'gestion/cambiar_contrasena.html')

@login_required
def descargar_historial_excel(request):
    """
    Vista para descargar el historial de caja filtrado en formato Excel
    """
    # Obtener los mismos filtros que en historial_caja
    registros = Registro.objects.filter(estado='abonado').order_by('-fecha')
    producto_nombre = request.GET.get('producto')
    doctor_id = request.GET.get('doctor')
    fecha_inicio = request.GET.get('desde')
    fecha_fin = request.GET.get('hasta')
    caja_id = request.GET.get('caja')
    
    # Aplicar los mismos filtros
    if producto_nombre:
        registros = registros.filter(productoutilizado__producto__nombre=producto_nombre)
    if doctor_id:
        registros = registros.filter(doctor__id=doctor_id)
    if fecha_inicio:
        registros = registros.filter(fecha__gte=parse_date(fecha_inicio))
    if fecha_fin:
        registros = registros.filter(fecha__lte=parse_date(fecha_fin))
    if caja_id:
        registros = registros.filter(caja_destino__id=caja_id)
    
    # Crear el libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Historial de Caja"
    
    # Estilos para el header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="007bff", end_color="007bff", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers de las columnas
    headers = ['Fecha', 'Paciente', 'Tratamiento', 'Consultorio', 'Doctor', 'Productos', 'Abonado en', 'Estado']
    
    # Agregar headers con estilo
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Agregar datos
    row = 2
    for registro in registros:
        # Obtener nombre completo del doctor
        try:
            perfil = PerfilUsuario.objects.get(usuario=registro.doctor)
            doctor_nombre = perfil.nombre_completo()
        except PerfilUsuario.DoesNotExist:
            doctor_nombre = registro.doctor.get_full_name() if registro.doctor.get_full_name() else registro.doctor.username
        
        # Obtener lista de productos
        productos_list = []
        for item in registro.productoutilizado_set.all():
            productos_list.append(f"{item.producto.nombre} ({item.cantidad})")
        productos_texto = ", ".join(productos_list)
        
        # Agregar fila de datos
        ws.cell(row=row, column=1, value=registro.fecha.strftime("%d/%m/%Y"))
        ws.cell(row=row, column=2, value=registro.nombre_paciente)
        ws.cell(row=row, column=3, value=registro.tratamiento.nombre)
        ws.cell(row=row, column=4, value=registro.consultorio.nombre)
        ws.cell(row=row, column=5, value=doctor_nombre)
        ws.cell(row=row, column=6, value=productos_texto)
        ws.cell(row=row, column=7, value=registro.caja_destino.nombre)
        ws.cell(row=row, column=8, value="Abonado")
        
        # Aplicar bordes a las celdas de datos
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = border
        
        row += 1
    
    # Ajustar ancho de columnas
    column_widths = [12, 25, 20, 15, 25, 40, 20, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # Crear nombre del archivo con timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"historial_caja_{timestamp}.xlsx"
    
    # Preparar response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Guardar el archivo en el response
    wb.save(response)
    
    # Log de la descarga
    logger.info(f"Usuario {request.user.username} descargó historial Excel: {filename}")
    
    return response
